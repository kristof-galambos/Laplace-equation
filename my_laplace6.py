"""
This code solves the 2D Laplace equation subject to given boundary conditions.

The temperature of the room is shown in animated figure.
The input is from an excel file "laplace_in3.xlsx"
The speed of the animation is inversely proportional to the grid size squared.
"""
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.animation as ani
import openpyxl
import sys

#read in input details from excel file:
book = openpyxl.load_workbook('laplace_in3.xlsx')
sheet = book.active

try:
    room_temp = float(sheet['D2'].value)
    grid_width = int(sheet['D3'].value)
    grid_height = int(sheet['D4'].value)
except:
    sys.exit('Invalid input details, check "laplace_in3.xlsx"!')

eqn = np.zeros([grid_height, grid_width])+room_temp
for i in range(grid_width): #top wall
    eqn[0][i] = float(sheet.cell(row=7+i, column=1).value)
for i in range(grid_width): #bottom wall
    eqn[grid_height-1][i] = float(sheet.cell(row=7+i, column=2).value)
for i in range(grid_height): #left wall
    eqn[i][0] = float(sheet.cell(row=7+i, column=3).value)
for i in range(grid_height): #right wall
    eqn[i][grid_width-1] = float(sheet.cell(row=7+i, column=4).value)

#add a wierd blob of hotter air in the room. OPTIONAL
for row in range(25,50):
    for col in range(40,60):
        eqn[row][col] = 50


fig = plt.figure(1)
ax = fig.add_subplot(111)

def init():
    im = ax.imshow(eqn, cmap='RdBu')
    ax.set_title(r'Temperature in the room ($^\circ$C)')
    ax.set_xlabel('Dimensions of the walls of the room are metres')
    return im,

def animate(t):
    i =1
    while i<grid_height-1:
        j =1
        while j<grid_width-1:
            eqn[i][j] = 0.25*(eqn[i-1][j] + eqn[i+1][j] + eqn[i][j-1] + eqn[i][j+1])
            j+=1
        i+=1
    im = ax.imshow(eqn, cmap='RdBu')
    return im,

anim = ani.FuncAnimation(fig, animate, init_func=init, frames=5000, interval=10, blit=True)